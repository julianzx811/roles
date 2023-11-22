import random
from django.urls import reverse
import smtplib
import openpyxl
import smtplib
from django.forms.models import model_to_dict
from django.http import FileResponse, HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
import plotly.express as px

from .forms import (DatosForm, FileUploadARLForm, FileUploadEPSForm,
                    FileUploadLABORALForm, ProgramForm)
from .models import (Aspirantes, Contrato, Coordinador, Estado_Practica,
                     Estudiante, Perfiles, Plan_estudios, Programas, Semestres,
                     UploadedARLFile, UploadedEPSFile, UploadedLABORALFile,
                     monitores)

archivo_subido = True


def login(request):
    if request.method == "GET":
        return render(request, "Archivos/login.html")
    if request.method == "POST":
        login_fallo = False
        usuario_existe = False
        usuario = request.POST["usuario"]
        contrasena = request.POST["contrasena"]
        cargo = request.POST["cargoUsuario"]
        print(usuario, contrasena, cargo)
        if Perfiles.objects.filter(usuario=usuario).exists():
            if (
                (Perfiles.objects.get(usuario=usuario).usuario == usuario)
                and (Perfiles.objects.get(usuario=usuario).contrasena == contrasena)
                and (Perfiles.objects.get(usuario=usuario).cargo == cargo)
            ):
                print("Login exitoso")
                if cargo == "Practicante":
                    return redirect("/vistaEstudiante/" + str(usuario))
                elif cargo == "Docente Monitor":
                    return redirect(
                        "/vistaDocenteMonitor",
                    )
                elif cargo == "Coordinador":
                    return redirect("/vistaCoordinador")
                elif cargo == "Lider Oficina de Practicas":
                    return redirect("/vistaLiderOficinaPracticas")
                elif cargo == "Auxiliar Oficina de Practicas":
                    return redirect("/vistaAuxiliarOficinaPracticas")
                elif cargo == "Administrador":
                    return redirect("/vistaAdministrador")
            else:
                login_fallo = True
                return render(
                    request,
                    "Archivos/login.html",
                    {"login_fallo": login_fallo, "usuario_existe": usuario_existe},
                )
        else:
            usuario_existe = True
            return render(
                request,
                "Archivos/login.html",
                {"login_fallo": login_fallo, "usuario_existe": usuario_existe},
            )


def recuperarContrasena(request):
    if request.method == "GET":
        return render(request, "Archivos/olvideContrasena.html")
    if request.method == "POST":
        usuario_existe = False
        correo = request.POST["usuario"]
        if Perfiles.objects.filter(correo=correo).exists():
            print("Hola, si existes")
            usuario = correo.split("@")[0]
            numeroVerificacion = random.randrange(100000,999999)
            email_usuario = Perfiles.objects.filter(usuario=usuario)[0].correo
            smtp_server = 'smtp.office365.com'
            smtp_port = 587

            sender = 'smartinezs2@correo.usbcali.edu.co'
            contraseña = 'Usb30000062842'


            receivers = [email_usuario]
            message = """\
            Subject: Codigo de Verificacion

            Hola, soy el sistema de seguimiento de practicas. Has solicitado un cambio de contrasena, el numero de verificacion es""" + str(numeroVerificacion)+ """.
            Espero haberte ayudado
            """

            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(sender, contraseña)
            server.sendmail(sender, receivers, message)
            server.quit()
            return redirect(verificarNumero, usuario = correo.split("@")[0], numeroVerificacion=numeroVerificacion)
        else:
            usuario_existe = True
            return render(
                request,
                "Archivos/olvideContrasena.html",
                {"usuario_existe":usuario_existe},
            )


def verificarNumero(request, usuario, numeroVerificacion):

    if request.method == 'POST':
        ingresado = request.POST['numeroVerificacionIngresado']
        print(ingresado,numeroVerificacion)
        if ingresado == numeroVerificacion:
            return redirect(ingresarNuevaContrasena, usuario = usuario)
        else:
            print("Intenta de nuevo")
            autenticacion_fallo = True
            return render(
                request,
                "Archivos/ingresarNumeroVerificacion.html",{"autenticacion_fallo":autenticacion_fallo}
            )
    return render(
                request,
                "Archivos/ingresarNumeroVerificacion.html",
    )
        
    

def ingresarNuevaContrasena(request, usuario):
    if request.method == 'POST':
        contrasenaNueva = request.POST['contrasena1']
        contrasenaVerificar = request.POST['contrasena2']
        if contrasenaNueva == contrasenaVerificar:
            contrasenas_coinciden = True
            Perfiles.objects.filter(usuario=usuario).update(contrasena = contrasenaNueva)
            return render(
                request,
                "Archivos/ingresarNuevaContrasena.html",{"contrasenas_coinciden":contrasenas_coinciden}
            )
        else:
            contrasenas_coinciden = False
            return render(
                request,
                "Archivos/ingresarNuevaContrasena.html",{"contrasenas_coinciden":contrasenas_coinciden}
            )

    return render(
                request,
                "Archivos/ingresarNuevaContrasena.html",
            )

def legalizacionEstudiantes(request):
    if request.method == "GET":
        semestres = Semestres.objects.all()
        print(semestres)
        return render(
            request,
            "Archivos/legalizacionEstudiantes.html", {"semestres":semestres}
        )
    elif request.method == "POST":
        periodo = request.POST['periodo_lectivo']
        if 'uno' in request.POST:    
            return redirect(aprobarLegalizacionEstudiantes, periodo = periodo)
    else:
        return render(
            request, "Archivos/legalizacionEstudiantes.html"
        )
    
def aprobarLegalizacionEstudiantes(request, periodo):
    if request.method == "GET":
        estudiantes = Estudiante.objects.filter(periodo_lectivo = periodo , estado_legalizacion = 'Pendiente')
        return render(
            request,
            "Archivos/aprobarLegalizacionEstudiantes.html", {'estudiantes':estudiantes}
        )
    elif request.method == "POST":
        
        codigo = request.POST['aprobarPractica']
        Estudiante.objects.filter(codigo = codigo).update(estado_legalizacion = 'Aprobado')

        return redirect(aprobarLegalizacionEstudiantes, periodo = periodo)
    else:
        return redirect(aprobarLegalizacionEstudiantes, periodo = periodo)

def AsignacionDocentesEstudiantes(request):
    mostrar = None
    if request.method == "GET":
        semestres = Semestres.objects.all()
        mostrar = None
        print(semestres)
        return render(
            request,
            "Archivos/asignacionDocentesEstudiantes.html",
            {"mostrar": mostrar, "semestres": semestres},
        )
    elif request.method == "POST":
        mostrar = None
        print(request.POST)
        mostrar = request.POST.get("periodo_lectivo")
        print(mostrar)
        docentes = monitores.objects.filter(estado=True)
        estudiantes = Estudiante.objects.filter(periodo_lectivo=mostrar)
        horas_totales = 0
        estudiantes_aleatorios = [x for x in estudiantes]
        random.shuffle(estudiantes_aleatorios)
        for docente in docentes:
            horas_totales += docente.horas_disponibles
        porcentajes_docentes = []
        for docente in docentes:
            porcentajes_docentes.append(
                round(
                    (len(estudiantes) * (docente.horas_disponibles / horas_totales)), 0
                )
            )
        if len(estudiantes) == 1:
            porcentajes_docentes[0] += 1
        acumulado = 0
        for idx, x in enumerate(porcentajes_docentes):
            for y in range(acumulado, acumulado + int(x)):
                estudiante = Estudiante.objects.filter(
                    codigo=estudiantes_aleatorios[acumulado].codigo
                ).update(docente_asignado_id=docentes[idx])
                acumulado += 1

        print(acumulado)

        return render(
            request,
            "Archivos/asignacionDocentesEstudiantes.html",
            {"mostrar": mostrar, "docentes": docentes, "estudiantes": estudiantes},
        )
    else:
        return render(
            request, "Archivos/asignacionDocentesEstudiantes.html", {"mostrar": mostrar}
        )


def CrearMonitor(request):
    programas = Programas.objects.filter()
    print(programas)
    if request.method == "GET":
        return render(request, "Archivos/CrearMonitor.html", {"programas": programas})
    else:
        creaMonitor = False
        monitorExiste = False
        form = DatosForm(request.POST)
        usuario = request.POST["correo"].split("@")[0]
        print(usuario)
        if form.is_valid():
            if Perfiles.objects.filter(usuario=usuario).exists() == False:
                Nombre = request.POST["nombre"]
                Codigo = request.POST["codigo"]
                Correo = request.POST["correo"]
                Horas = request.POST["horas"]
                Programa = request.POST.get("programa")
                programa_asignado = Programas.objects.get(programa=Programa)
                monitor = monitores(
                    nombre=Nombre,
                    codigo=Codigo,
                    correo_institucional=Correo,
                    horas_disponibles=Horas,
                    programa=programa_asignado,
                )
                monitor.save()
                perfil = Perfiles(
                    usuario=Correo.split("@")[0],
                    contrasena=Codigo,
                    nombre=Nombre,
                    cargo="Docente Monitor",
                    correo=Correo
                )
                perfil.save()
                creaMonitor = True
            else:
                monitorExiste = True
        return render(
            request,
            "Archivos/CrearMonitor.html",
            {
                "archivo_subido": archivo_subido,
                "creaMonitor": creaMonitor,
                "monitorExiste": monitorExiste,
                "programas": programas,
            },
        )


def crearPrograma(request):
    return render(request, "Archivos/CrudPrograma.html")


def index(request):
    if request.method == "GET":
        return render(request, "Archivos/index.html")


def indexCoordinador(request):
    existe = Estudiante.objects.exists()
    return render(
        request,
        "Archivos/vistaCoordinador.html",
        {"archivo_subido": archivo_subido, "existe": existe},
    )


def indexOficinaPracticas(request):
    existe = Estudiante.objects.all()
    estudiantesPendientes = Estudiante.objects.filter(estado_legalizacion = 'Pendiente')
    estudiantesAprobado = Estudiante.objects.filter(estado_legalizacion = 'Aprobado')
    estudiantesIncompleto = Estudiante.objects.filter(estado_legalizacion = 'Incompleto')
    print(estudiantesPendientes)
    data = [len(estudiantesPendientes), len(estudiantesAprobado), len(estudiantesIncompleto)]
    labels = ['Pendientes', 'Aprobados', 'Incompletos']

    fig = px.pie(
        names=labels, 
        values=data, 
        title='Estado de Legalización de Estudiantes'  # Ajusta la opacidad según sea necesar
        )
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
         font_color='white'
    )

    # Convierte la figura de Plotly a HTML
    plot_div = fig.to_html(full_html=False)
    return render(
        request,
        "Archivos/vistaLiderOficinaPracticas.html",
        {"archivo_subido": archivo_subido, "existe": existe,"plot_div": plot_div,"estudiantesAprobado":estudiantesAprobado,"estudiantesPendientes":estudiantesPendientes,"estudiantesIncompleto":estudiantesIncompleto},
    )


def indexAuxiliarOficinaPracticas(request):
    existe = Estudiante.objects.exists()
    return render(
        request,
        "Archivos/vistaAuxiliarOficinaPracticas.html",
        {"archivo_subido": archivo_subido, "existe": existe},
    )


def indexDocenteMonitor(request):
    existe = Estudiante.objects.exists()
    return render(
        request,
        "Archivos/index.html",
        "Archivos/vistaDocenteMonitor.html",
        {"archivo_subido": archivo_subido, "existe": existe},
    )


def indexEstudiante(request, estudiante_id):
    existe = Estudiante.objects.exists()
    return render(
        request,
        "Archivos/vistaEstudiante.html",
        {
            "archivo_subido": archivo_subido,
            "existe": existe,
            "estudiante_id": estudiante_id,
        },
    )


def indexAdministrador(request):
    existe = Estudiante.objects.exists()
    return render(
        request,
        "Archivos/vistaAdministrador.html",
        {
            "archivo_subido": archivo_subido,
            "existe": existe,
        },
    )


def agregarNuevoLider(request):
    if request.method == "GET":
        return render(request, "Archivos/CrearLiderOficina.html")
    else:
        creaLiderOficina = False
        liderExiste = False

        usuario = request.POST["correo"].split("@")[0]

        if Perfiles.objects.filter(usuario=usuario).exists() == False:
            Nombre = request.POST["nombre"]
            Codigo = request.POST["codigo"]
            Correo = request.POST["correo"]
            perfil = Perfiles(
                usuario=Correo.split("@")[0],
                contrasena=Codigo,
                nombre=Nombre,
                cargo="Lider Oficina de Practicas",
                correo=Correo
            )
            perfil.save()
            creaLiderOficina = True
        else:
            liderExiste = True
        return render(
            request,
            "Archivos/CrearLiderOficina.html",
            {"creaLiderOficina": creaLiderOficina, "liderExiste": liderExiste},
        )


def agregarNuevoAuxiliar(request):
    if request.method == "GET":
        return render(request, "Archivos/CrearAuxiliarOficina.html")
    else:
        creaAuxiliarOficina = False
        auxiliarExiste = False

        usuario = request.POST["correo"].split("@")[0]

        if Perfiles.objects.filter(usuario=usuario).exists() == False:
            Nombre = request.POST["nombre"]
            Codigo = request.POST["codigo"]
            Correo = request.POST["correo"]
            perfil = Perfiles(
                usuario=Correo.split("@")[0],
                contrasena=Codigo,
                nombre=Nombre,
                cargo="Lider Oficina de Practicas",
                correo=Correo
            )
            perfil.save()
            creaAuxiliarOficina = True
        else:
            auxiliarExiste = True
        return render(
            request,
            "Archivos/CrearAuxiliarOficina.html",
            {
                "creaAuxiliarOficina": creaAuxiliarOficina,
                "auxiliarExiste": auxiliarExiste,
            },
        )


def asignarNuevoCoordinador(request):
    docentes = monitores.objects.filter()
    creaMonitor = False
    monitorExiste = False
    if request.method == "GET":
        return render(
            request, "Archivos/asignarNuevoCoordinador.html", {"docentes": docentes}
        )
    else:
        creaMonitor = False
        monitorExiste = False
        coordinador = monitores.objects.filter(codigo=request.POST["docente"])
        correo = coordinador[0].correo_institucional
        if Coordinador.objects.filter(id_docente=correo).exists() == False:
            monitor = monitores.objects.filter(codigo=request.POST["docente"])[0]
            coordinador = Coordinador(id_docente=monitor)
            coordinador.save()

            perfil = Perfiles(
                usuario=correo.split("@")[0],
                contrasena=monitor.codigo,
                nombre=monitor.nombre,
                cargo="Coordinador",
                correo=correo
            )
            perfil.save()
            creaMonitor = True
        else:
            monitorExiste = True
        return render(
            request,
            "Archivos/asignarNuevoCoordinador.html",
            {
                "archivo_subido": archivo_subido,
                "creaMonitor": creaMonitor,
                "monitorExiste": monitorExiste,
                "docentes": docentes,
            },
        )


def cargarArchivoEstudiantes(request):
    if request.method == "GET":
        semestre = Semestres.objects.all()
        return render(request, "Archivos/cargaEstudiantes.html", {"semestre": semestre})
    else:
        try:
            actualizados = 0
            agregados = 0
            excel_file = request.FILES["excel_file"]
            periodo = request.POST.get("file_type")
            semestre_seleccionado = Semestres.objects.get(id=periodo)
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb["Sheet1"]

            excel_data = []
            for row in worksheet.iter_rows(min_row=14, max_col=8):
                row_data = []
                for cell in row:
                    if str(cell.value) != "None":
                        row_data.append(str(cell.value))
                if (
                    len(row_data) >= 7
                ):  # Verificar si hay suficientes elementos en la lista
                    excel_data.append(row_data)
                else:
                    print(f"La fila no tiene suficientes elementos: {row_data}")

            for datos in excel_data:
                if Estudiante.objects.filter(codigo=datos[2]):
                    if (
                        (
                            (Estudiante.objects.get(codigo=datos[2]).nombre == datos[6])
                            == False
                        )
                        or (
                            (
                                Estudiante.objects.get(
                                    codigo=datos[2]
                                ).email_institucional
                                == datos[3]
                            )
                            == False
                        )
                        or (
                            (
                                Estudiante.objects.get(codigo=datos[2]).email_personal
                                == datos[4]
                            )
                            == False
                        )
                        or (
                            (
                                Estudiante.objects.get(codigo=datos[2]).telefono
                                == datos[5]
                            )
                            == False
                        )
                    ):
                        actualizados += 1
                        print(actualizados)
                        Estudiante.objects.filter(codigo=datos[2]).update(
                            email_institucional=datos[3],
                            email_personal=datos[4],
                            telefono=datos[5],
                            nombre=datos[6],
                        )
                        perfil = Perfiles(
                        usuario=datos[3].split("@")[0],
                        contrasena=datos[2],
                        nombre=datos[6],
                        cargo="Practicante",
                        correo=datos[3]
                        )
                        perfil.save()
                        
                else:
                    est1 = Estudiante(
                        codigo=datos[2],
                        programa=datos[1],
                        email_institucional=datos[3],
                        email_personal=datos[4],
                        telefono=datos[5],
                        nombre=datos[6],
                        periodo_lectivo=semestre_seleccionado.nombre,
                    )
                    agregados += 1
                    print("agregados", agregados)
                    est1.save()
                    perfil = Perfiles(
                        usuario=datos[3].split("@")[0],
                        contrasena=datos[2],
                        nombre=datos[6],
                        cargo="Practicante",
                        correo=datos[3]
                    )
                    perfil.save()
            return render(
                request,
                "Archivos/cargaEstudiantes.html",
                {
                    "excel_data": excel_data,
                    "agregados": agregados,
                    "actualizados": actualizados,
                },
            )

        except Exception as error:
            print(error)
            return render(
                request, "Archivos/CargaEstudiantes.html", {"error": str(error)}
            )


def cargarArchivoEstudiantesDos(request):
    existe = Estudiante.objects.exists()
    semestre_no_existe = False
    if request.method == "GET":
        semestre = Semestres.objects.all()
        return render(
            request, "Archivos/CargaEstudiantesDos.html", {"semestre": semestre}
        )
    else:
        try:
            periodo = request.POST.get("file_type")
            semestre_seleccionado = Semestres.objects.get(id=periodo)
            print(
                len(
                    Estudiante.objects.filter(
                        periodo_lectivo=semestre_seleccionado.nombre
                    )
                )
            )
            if (
                len(
                    Estudiante.objects.filter(
                        periodo_lectivo=semestre_seleccionado.nombre
                    )
                )
                != 0
            ):
                actualizados = 0
                excel_file = request.FILES["excel_file"]
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb["ING SISTEMAS 2023 2"]
                excel_data = []

                for row in worksheet.iter_rows(min_row=4, max_col=28):
                    row_data = []
                    for cell in row:
                        if str(cell.value) != "None":
                            row_data.append(str(cell.value))
                    if (
                        len(row_data) >= 7
                    ):  # Verificar si hay suficientes elementos en la lista
                        excel_data.append(row_data)

                for row in excel_data:
                    # Verifica si la fila tiene suficientes elementos para procesar
                    if (
                        len(row) >= 28
                    ):  # Ajusta el número según la cantidad de columnas en tu Excel
                        # Crear un objeto Estudiante

                        plan_estudios = Plan_estudios(jornada=row[12])
                        plan_estudios.save()

                        if Estudiante.objects.filter(codigo=row[7]):
                            print("estudiante ya existe")
                            estudiante = Estudiante.objects.filter(codigo=row[7])[0]
                            Estudiante.objects.filter(codigo=row[7]).update(
                                nombre=row[5],
                                apellidos=row[6],
                                cedula=row[8],
                                celular=row[9],
                            )

                            estudiante = Estudiante(
                                programa=row[11],
                                codigo=row[7],
                                email_institucional=row[10],
                                email_personal=row[10],
                                telefono=row[9],
                                nombre=row[5],
                                apellidos=row[6],
                                cedula=row[8],
                                celular=row[9],
                                periodo_lectivo=row[1]
                                # Ajusta esto para el campo plan_estudios
                            )
                            actualizados += 1
                            estudiante.save()

                            # Crear un objeto Aspirantes relacionado con el estudiante
                            aspirantes = Aspirantes(
                                periodo_practica=row[1],
                                aprobación_Programa=row[2],
                                matriculado_Academica_y_Financieramente=row[4],
                                inscripcion=row[13],
                                curso_induccion_y_rl=row[14],
                                ruta_preparacion_vida_laboral=row[15],
                                envio_hv=row[16],
                                titulo_tecnico_o_tecnologo=row[17],
                                codigo_estudiante=estudiante,
                            )
                            aspirantes.save()

                            if "Aplaza" in row[1]:
                                periodo_aplazado = row[1].split("Aplaza ")[1]
                                periodo_aplazado = periodo_aplazado[:-1]
                                estudiante.periodo_lectivo = periodo_aplazado
                                print(
                                    "Periodo despues de aplazar: "
                                    + row[1].split("Aplaza ")[1]
                                )
                            elif row[1] == "NO APROBADO":
                                estudiante.periodo_lectivo = "suspendido"

                            estudiante.save()

                            # Crear un objeto Contrato relacionado con el estudiante
                            print("entro If")
                            contrato = Contrato(
                                tipo_Contrato=row[21],
                                fecha_Inicio=row[22],
                                fecha_Final=row[23],
                                encargado_Proceso_Seleccion=row[24],
                                datos_Tutor_O_Jefe_Directivo=row[25],
                                documentos_Pendientes=row[26],
                                sector=row[27],
                            )
                            contrato.save()

                            # Crear un objeto Estado_Practica relacionado con el estudiante, Aspirantes y Contrato
                            estado_practica = Estado_Practica(
                                codigo_estudiante=estudiante,
                                practica_Donde_Labora_EmpresaFliar_Emprendim_Otro=row[
                                    18
                                ],
                                estado_ubicación=row[19],
                                comentarios=row[20],
                                item=aspirantes,
                                id_contrato=contrato,
                            )
                            estado_practica.save()
                            return render(
                                request,
                                "Archivos/CargaEstudiantesDos.html",
                                {
                                    "semestre_no_existe": semestre_no_existe,
                                    "excel_data": excel_data,
                                    "existe": existe,
                                    "actualizados": actualizados,
                                },
                            )
            else:
                semestre_no_existe = True
                return render(
                    request,
                    "Archivos/CargaEstudiantesDos.html",
                    {"semestre_no_existe": semestre_no_existe},
                )
            return render(
                request,
                "Archivos/CargaEstudiantesDos.html",
                {
                    "semestre_no_existe": semestre_no_existe,
                    "excel_data": excel_data,
                    "existe": existe,
                    "actualizados": actualizados,
                },
            )
        except Exception as error:
            print(error)
            return render(
                request, "Archivos/CargaEstudiantesDos.html", {"error": str(error)}
            )


def MostrarEstudiantes(request):
    estudiantes = Estudiante.objects.filter()
    return render(request, "Archivos/verEstudiantes.html", {"estudiantes": estudiantes})


def CrudPrograma(request):
    programas = Programas.objects.all()
    if request.method == "DELETE":
        return render(request, "Archivos/CrudPrograma.html", {"programas": programas})
    if request.method == "GET":
        return render(request, "Archivos/CrudPrograma.html", {"programas": programas})
    if request.method == "POST":
        form = ProgramForm(request.POST)
        if form.is_valid():
            try:
                codigo = form.cleaned_data["codigo"]
                programa = form.cleaned_data["programa"]
                facultad = request.POST.get("facultad")
                programita = Programas(
                    codigo=codigo, programa=programa, facultad=facultad
                )
                programita.save()
                return render(
                    request,
                    "Archivos/CrudPrograma.html",
                    {
                        "programas": programas,
                        "hecho": "se registro el monitor correctamente",
                    },
                )
            except Exception as error:
                print(error)
                return render(
                    request,
                    "Archivos/CrudPrograma.html",
                    {"error": error, "programas": programas},
                )
        else:
            print(form.errors)
            return render(
                request,
                "Archivos/CrearMonitor.html",
                {"error": "algo salio mal", "programas": programas},
            )


def UpdatePrograma(request, programa_id):
    if request.method == "GET":
        programa = Programas.objects.get(pk=programa_id)
        programaobj = model_to_dict(programa)
        return render(
            request,
            "Archivos/UpdatePrograma.html",
            {"programa": programaobj, "id": programa_id},
        )
    elif request.method == "POST":
        print("entro1")
        form = ProgramForm(request.POST)
        if form.is_valid():
            try:
                print("entro2")
                codigo = form.cleaned_data["codigo"]
                programa = form.cleaned_data["programa"]
                facultad = request.POST.get("facultad")
                Programas.objects.filter(id=programa_id).update(
                    codigo=codigo, programa=programa, facultad=facultad
                )
                request.method = "GET"
                return CrudPrograma(request)
            except Exception as error:
                print(error)
                return CrudPrograma(request)


def CreatePrograma(request):
    return render(request, "Archivos/CreatePrograma.html")


def DeletePrograma(request, programa_id):
    programa = Programas.objects.get(pk=programa_id)
    try:
        programa.delete()
        return CrudPrograma(request)
    except Exception as error:
        print(error)
        return CrudPrograma(request)


def list_files(request, estudiante_id):
    try:
        files1 = UploadedARLFile.objects.filter(estudianteId=estudiante_id)
        files2 = UploadedEPSFile.objects.filter(estudianteId=estudiante_id)
        files3 = UploadedLABORALFile.objects.filter(estudianteId=estudiante_id)
        context = {
            "files1": files1,
            "files2": files2,
            "files3": files3,
        }
        return render(request, "list_files.html", context)
    except Exception as error:
        print(error)
        return render(
            request, "list_files.html", {"files1": None, "files2": None, "files3": None}
        )


def view_file(request, file_id):
    file = get_object_or_404(FileUploadARLForm, id=file_id)
    return FileResponse(open(file.file.path, "rb"), as_attachment=True)


def iniciar_practicas(request, estudiante_id):
    return render(request, "Archivos/iniciarPracticas.html")


def upload_file(request, stringhtml, formxd, estudiante_id):
    perfil = Perfiles.objects.get(pk=estudiante_id)
    print(perfil)
    if request.method == "POST":
        form = formxd(request.POST, request.FILES)
        if form.is_valid():
            file = UploadedLABORALFile.objects.filter(estudianteId=perfil)
            if file.exists():
                existing_file = (
                    file.first()
                )  # Assuming there's only one file per student
                form = formxd(request.POST, request.FILES, instance=existing_file)
                if form.is_valid():
                    # Update the existing file with the new data
                    form.save()
                    return redirect("../")

            else:
                uploaded_file = form.save(commit=False)
                uploaded_file.estudianteId = perfil
                uploaded_file.save()
                return redirect("../")
    else:
        form = formxd()
    return render(request, str(stringhtml), {"form": form})


def SubirContranoLaboral(request, estudiante_id):
    return upload_file(
        request, "SubirContranoLaboral.html", FileUploadLABORALForm, estudiante_id
    )


def SubirAfiliacionARL(request, estudiante_id):
    return upload_file(
        request, "SubirAfiliacionARL.html", FileUploadARLForm, estudiante_id
    )


def SubirDocumentoEPS(request, estudiante_id):
    return upload_file(
        request, "SubirDocumentoEPS.html", FileUploadEPSForm, estudiante_id
    )


def DocumentosSubidos(request, estudiante_id):
    if len(UploadedEPSFile.objects.filter(estudianteId_id = estudiante_id)) > 0 and len(UploadedLABORALFile.objects.filter(estudianteId_id = estudiante_id)) > 0:
        estudiante = Estudiante.objects.filter(email_institucional = Perfiles.objects.filter(usuario = estudiante_id)[0].correo)
        estudiante.update(estado_legalizacion = 'Pendiente')
    
    return list_files(request, estudiante_id)


def administrarSemestres(request):
    semestres = Semestres.objects.all()
    if request.method == "DELETE":
        return render(
            request, "Archivos/administrarSemestre.html", {"semestres": semestres}
        )
    if request.method == "GET":
        return render(
            request, "Archivos/administrarSemestre.html", {"semestres": semestres}
        )
    if request.method == "POST":
        try:
            semestre = request.POST["semestre"]
            fecha_inicio = request.POST["fecha_inicio"]
            fecha_fin = request.POST["fecha_fin"]
            semestrito = Semestres(
                nombre=semestre, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin
            )
            print(semestrito)
            semestrito.save()
            return render(
                request,
                "Archivos/administrarSemestres.html",
                {
                    "semestres": semestres,
                },
            )
        except Exception as error:
            print(error)
            return render(
                request,
                "Archivos/administrarSemestre.html",
                {"semestres": semestres},
            )


def UpdateSemestre(request, id):
    if request.method == "GET":
        semestre = Semestres.objects.get(pk=id)
        semestreobj = model_to_dict(semestre)
         
        # Formatea la fecha en el formato "AAAA-MM-DD"
        fecha_inicio_formateada = semestreobj['fecha_inicio'].strftime("%Y-%m-%d")
        semestreobj['fecha_inicio'] = fecha_inicio_formateada
        fecha_fin_formateada = semestreobj['fecha_fin'].strftime("%Y-%m-%d")
        semestreobj['fecha_fin'] = fecha_fin_formateada
        
        return render(
            request,
            "Archivos/UpdateSemestre.html",
            {"semestre": semestreobj, "id": id},
        )
    elif request.method == "POST":
        try:
            
            semestre = request.POST["semestre"]
            fecha_inicio = request.POST["fecha_inicio"]
            fecha_fin = request.POST["fecha_fin"]
            aux = Semestres.objects.filter(id=id).update(
                nombre=semestre, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin
            ) 
            request.method = "GET"
            return administrarSemestres(request)
        except Exception as error:
            print("eror: ",error)
            return administrarSemestres(request)


def CreateSemestre(request):
    return render(request, "Archivos/CreateSemestre.html")


def DeleteSemestre(request, id):
    semestre = Semestres.objects.get(pk=id)
    try:
        semestre.delete()
        return administrarSemestres(request)
    except Exception as error:
        print(error)
        return administrarSemestres(request)

def adminVisualizarEstudiantes(request):
    estudiantes = Estudiante.objects.all()      
    
    return render(request, "Archivos/listadoCrudEstudiantes.html", {"estudiantes": estudiantes})


def UpdateEstudiante(request, codigo):
    estudiante = get_object_or_404(Estudiante, codigo=codigo)
    print(estudiante)
    if request.method == 'POST':
        try:
            # Actualiza los campos directamente desde el request.POST
            estudiante.programa = request.POST['programa']
            estudiante.codigo = request.POST['codigo']
            estudiante.email_institucional = request.POST['email_institucional']
            estudiante.email_personal = request.POST['email_personal']
            estudiante.telefono = request.POST['telefono']
            estudiante.nombre = request.POST['nombre']
            estudiante.apellidos = request.POST['apellidos']
            estudiante.cedula = request.POST['cedula']
            estudiante.celular = request.POST['celular']
            estudiante.periodo_lectivo = request.POST['periodo_lectivo']
            estudiante.plan_estudios = Plan_estudios.objects.get(id=int(request.POST['plan_estudios_id']))
            estudiante.docente_asignado = monitores.objects.get(correo_institucional=request.POST['docente_asignado_id'])

            
            estudiante.save()

            return redirect('adminVisualizarEstudiantes')
        except Exception as e:
            # Manejo de errores, puedes personalizar según tus necesidades
            return HttpResponseBadRequest("Error en la actualización. Detalles: " + str(e))

    return render(request, 'Archivos/UpdateEstudiante.html', {'estudiante': estudiante})

def visualizarAdministrador(request):
    # Filtra los usuarios con el cargo "Administrador"
    administradores = Perfiles.objects.filter(cargo='Administrador')

    # Puedes enviar la lista de administradores a tu template
    return render(request, 'listadoCrudAdministrador.html', {'administradores': administradores})


def UpdateAdministrador(request, usuario):
    administrador = get_object_or_404(Perfiles, usuario=usuario, cargo='Administrador')

    if request.method == 'POST':
        try:
            # Actualiza los campos directamente desde el request.POST
            administrador.contrasena = request.POST['contrasena']
            administrador.usuario = request.POST['usuario']
            administrador.nombre = request.POST['nombre']
            administrador.cargo = request.POST['cargo']
            administrador.correo = request.POST['correo']

            administrador.save()

            return redirect('visualizarAdministrador')
        except Exception as e:
            # Manejo de errores, puedes personalizar según tus necesidades
            return HttpResponseBadRequest("Error en la actualización. Detalles: " + str(e))

    return render(request, 'Archivos/UpdateAdministrador.html', {'administrador': administrador})


def DeleteAdministrador(request, usuario):
    administrador = get_object_or_404(Perfiles, usuario=usuario)
    try:
        administrador.delete()
        return visualizarAdministrador(request)
    except Exception as error:
        print(error)
        return visualizarAdministrador(request)
    

def visualizarMonitor(request):
    _monitores = monitores.objects.all()  

    return render(request, 'listadoCrudMonitor.html', {'monitores': _monitores})


def UpdateMonitor(request, correo):
    monitor = get_object_or_404(monitores, correo_institucional=correo)

    if request.method == 'POST':
        try:
            # Actualiza los campos directamente desde el request.POST
              
            monitor.nombre = request.POST['nombre']
            monitor.codigo = request.POST['codigo']
            monitor.correo_institucional = request.POST['correo_institucional']
            monitor.horas_disponibles = request.POST['horas_disponibles']
            Programa = request.POST.get("programa")
            programa_instance = Programas.objects.get(programa=Programa)
            monitor.programa = programa_instance
            monitor.estado = request.POST['estado']
            

            monitor.save()

            return redirect('visualizarMonitor')
        except Exception as e:
            # Manejo de errores, puedes personalizar según tus necesidades
            return HttpResponseBadRequest("Error en la actualización. Detalles: " + str(e))

    return render(request, 'Archivos/UpdateMonitor.html', {'monitor': monitor})


def DeleteMonitor(request, correo):
    monitor = get_object_or_404(monitores, correo_institucional=correo)
    try:
        monitor.delete()
        return visualizarMonitor(request)
    except Exception as error:
        print(error)
        return visualizarMonitor(request)
    

def visualizarCoordinador(request):
    coordinadores = Coordinador.objects.select_related('id_docente').all()
    return render(request, 'Archivos/listadoCrudCoordinador.html', {'coordinadores': coordinadores})

def visualizarLiderOficinaPracticas(request):
    # Filtra los usuarios con el cargo "Administrador"
    lideres = Perfiles.objects.filter(cargo='Lider Oficina de Practicas')

    # Puedes enviar la lista de administradores a tu template
    return render(request, 'listadoCrudLiderOficinaPracticas.html', {'lideres': lideres})


def UpdateLider(request, usuario):
    lider = get_object_or_404(Perfiles, usuario=usuario, cargo='Lider Oficina de Practicas')

    if request.method == 'POST':
        try:
            # Actualiza los campos directamente desde el request.POST
            lider.contrasena = request.POST['contrasena']
            lider.usuario = request.POST['usuario']
            lider.nombre = request.POST['nombre']
            lider.cargo = request.POST['cargo']
            lider.correo = request.POST['correo']

            lider.save()

            return redirect('visualizarLiderOficinaPracticas')
        except Exception as e:
            # Manejo de errores, puedes personalizar según tus necesidades
            return HttpResponseBadRequest("Error en la actualización. Detalles: " + str(e))

    return render(request, 'Archivos/UpdateAdministrador.html', {'lider': lider})


def DeleteLider(request, usuario):
    lider = get_object_or_404(Perfiles, usuario=usuario)
    try:
        lider.delete()
        return visualizarLiderOficinaPracticas(request)
    except Exception as error:
        print(error)
        return visualizarLiderOficinaPracticas(request)
    
